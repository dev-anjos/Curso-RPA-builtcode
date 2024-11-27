from openpyxl import load_workbook
from docx import Document
from docx.shared import Pt

nomeArquiov = "Alunos.xlsx"
planilhaAlunos = load_workbook(nomeArquiov)
print(planilhaAlunos.sheetnames)

sheet_selecionada = planilhaAlunos["Nomes"]


for linha in range(2, len(sheet_selecionada["A"]) + 1):

    arquivoWord = Document('Certificado1.docx')

    estilo = arquivoWord.styles["Normal"]
    nomeAluno = sheet_selecionada['A%s' % linha].value
    dia = sheet_selecionada['B%s' % linha].value
    mes = sheet_selecionada['C%s' % linha].value
    ano = sheet_selecionada['D%s' % linha].value
    curso = sheet_selecionada['E%s' % linha].value
    instrutor = sheet_selecionada['F%s' % linha].value

    frase = (f"O Aluno {nomeAluno} Concluiu com sucesso o curso de {curso}, "
             f"com a carga horária de 20 horas, promovido pela escola de Cursos Online em {dia} {mes} de {ano}")

    for paragrafo in arquivoWord.paragraphs:
        if "@nome" in paragrafo.text:
            paragrafo.text = frase

            font = estilo.font
            font.name = "Calibri (Corpo)"
            font.size = Pt(24)
        if "@instrutor" in paragrafo.text:
            paragrafo.text = instrutor
            font = estilo.font
            font.name = "Calibri (Corpo)"
            font.size = Pt(24)

    arquivoWord.save(f'certificados/{nomeAluno}.docx')